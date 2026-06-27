"""
format_outputs.py
------------------
Applies consistent professional formatting (Arial font, bold headers,
autofit-ish column widths, number formats) to the generated workbooks:
  - data/cleaned_orders.xlsx
  - outputs/data_quality_report.xlsx
  - outputs/pivot_summary.xlsx

Also sets workbook author metadata (creator / last-modified-by) so the
file properties reflect the actual author rather than the default
library name.

Run after clean_orders.py:
    python scripts/format_outputs.py
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

AUTHOR_NAME = "Anagh Upadhya"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME)

CURRENCY_COLS = {
    "unit_price", "sales", "cost", "profit",
    "calculated_sales", "calculated_profit",
    "total_calculated_sales", "total_calculated_profit",
}
PERCENT_COLS = {"cleaned_discount", "profit_margin", "avg_profit_margin"}
DATE_COLS = {"order_date", "ship_date"}


def format_workbook(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        headers = [c.value for c in ws[1]]

        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len(str(val)) if val is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

            if header in CURRENCY_COLS:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0.00;($#,##0.00);-"
            elif header in PERCENT_COLS:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"
            elif header in DATE_COLS:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"

        ws.freeze_panes = "A2"

    # Set author metadata (overwrites openpyxl's default "openpyxl" creator)
    props = wb.properties
    props.creator = AUTHOR_NAME
    props.lastModifiedBy = AUTHOR_NAME
    props.title = None
    props.description = None
    props.subject = None
    props.keywords = None
    props.category = None
    props.company = None

    wb.save(path)


if __name__ == "__main__":
    for f in (
        "data/cleaned_orders.xlsx",
        "outputs/data_quality_report.xlsx",
        "outputs/pivot_summary.xlsx",
    ):
        format_workbook(f)
        print(f"Formatted {f}")
